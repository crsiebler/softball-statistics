"""
Excel exporter for softball statistics.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from softball_statistics.calculators.stats_calculator import (
    calculate_batting_average,
    calculate_ops,
    calculate_slg,
)
from softball_statistics.interfaces import QueryRepository
from softball_statistics.models import PlayerStats

logger = logging.getLogger(__name__)


class ExcelExportError(Exception):
    """Raised when Excel export fails."""


class ExcelExporter:
    """Excel exporter implementing Exporter interface."""

    def __init__(self, query_repo: QueryRepository):
        self.query_repo = query_repo

    def export(
        self,
        data: Dict[str, Any],
        output_path: str,
        team_name: Optional[str] = None,
        season: Optional[str] = None,
        use_case=None,
    ) -> None:
        """Export data to Excel."""
        export_to_excel(data, output_path, self.query_repo, team_name, season, use_case)


def _abbreviate_team_name(team_name: str) -> str:
    """
    Abbreviate team name to 5 characters for Excel sheet names.

    - Multiple words: creates acronym from first letters
    - Single word: truncates/pads to optimal length
    - Deterministic: same input always produces same output
    """
    if not team_name:
        return "Unknown"

    words = team_name.replace("_", " ").split()

    if len(words) > 1:
        # Create acronym from first letters of words
        acronym = "".join(word[0].upper() for word in words if word)
        return acronym[:5]  # Cap at 5 chars (acronyms are usually shorter)
    else:
        # Single word: aim for 5 characters
        word = words[0] if words else team_name
        if len(word) <= 5:
            return word
        else:
            return word[:5]  # Truncate to 5 chars


def export_to_excel(
    stats_data: Dict[str, Any],
    output_path: str,
    query_repo: Optional[QueryRepository] = None,
    team_name: Optional[str] = None,
    season: Optional[str] = None,
    use_case=None,
) -> None:
    """
    Export softball statistics to an Excel file with multiple sheets.

    Args:
        stats_data: Dictionary containing league/team stats and metadata
        output_path: Path to save the Excel file

    Raises:
        ExcelExportError: If export fails
    """
    try:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Create legend sheet first
            _create_legend_sheet(writer)

            # Always create league summary sheet
            _create_league_summary_sheet(stats_data, writer, query_repo, use_case)

            # Create comprehensive player summary sheet (if query_repo available)
            if query_repo:
                _create_player_summary_sheet(query_repo, writer)

            # Team sheets - now with cumulative and per-season structure
            if team_name and use_case:
                # Create cumulative total sheet
                cumulative_stats = use_case.get_cumulative_team_stats(team_name)
                _create_cumulative_team_sheet(team_name, cumulative_stats, writer)

                # Get all seasons for this team, sorted chronologically
                seasons = _get_seasons_for_team(team_name, query_repo)
                for season_name in seasons:
                    # Create season total sheet
                    season_stats = use_case.execute(
                        _get_league_for_team_season(team_name, season_name, query_repo),
                        team_name,
                        season_name,
                    )
                    _create_season_total_sheet(
                        team_name, season_name, season_stats, writer
                    )

                    # Create per-game sheets
                    if season_stats.get("team_stats", {}).get(team_name):
                        team_id = _get_team_id(team_name, season_name, query_repo)
                        if team_id:
                            games_stats = use_case.get_team_games_stats(team_id)
                            for game_stat in games_stats:
                                player_stats = use_case.get_game_player_stats(
                                    game_stat["game_id"]
                                )
                                _create_per_game_sheet(
                                    team_name, game_stat, player_stats, writer
                                )
            elif use_case and query_repo:
                # Multi-team detailed export mode - create detailed sheets for all teams
                all_teams = _get_all_teams(query_repo)
                for team_name_key in all_teams:
                    # Create cumulative team sheet
                    cumulative_stats = use_case.get_cumulative_team_stats(team_name_key)
                    if cumulative_stats.get("players"):  # Only create if team has data
                        _create_cumulative_team_sheet(
                            team_name_key, cumulative_stats, writer
                        )

                        # Get all seasons for this team, sorted chronologically
                        seasons = _get_seasons_for_team(team_name_key, query_repo)
                        for season_name in seasons:
                            # Create season total sheet
                            season_stats = use_case.execute(
                                _get_league_for_team_season(
                                    team_name_key, season_name, query_repo
                                ),
                                team_name_key,
                                season_name,
                            )
                            _create_season_total_sheet(
                                team_name_key, season_name, season_stats, writer
                            )

                            # Create per-game sheets
                            if season_stats.get("team_stats", {}).get(team_name_key):
                                team_id = _get_team_id(
                                    team_name_key, season_name, query_repo
                                )
                                if team_id:
                                    games_stats = use_case.get_team_games_stats(team_id)
                                    for game_stat in games_stats:
                                        player_stats = use_case.get_game_player_stats(
                                            game_stat["game_id"]
                                        )
                                        _create_per_game_sheet(
                                            team_name_key,
                                            game_stat,
                                            player_stats,
                                            writer,
                                        )
            else:
                # Fallback to original team sheets
                for team_name_key, team_stats in stats_data.get(
                    "team_stats", {}
                ).items():
                    _create_team_sheet(team_name_key, team_stats, writer)

            # Individual player sheets (if requested)
            if stats_data.get("include_player_details", False):
                for team_name, team_stats in stats_data.get("team_stats", {}).items():
                    for player_stats in team_stats.get("players", []):
                        _create_player_sheet(player_stats, writer)

        logger.info(f"Successfully exported statistics to {output_path}")

    except Exception as e:
        raise ExcelExportError(f"Failed to export to Excel: {e}")


def _create_league_summary_sheet(
    stats_data: Dict[str, Any],
    writer: pd.ExcelWriter,
    query_repo: Optional[QueryRepository] = None,
    use_case=None,
) -> None:
    """Create the league summary sheet with comprehensive database data."""

    # Try database-driven approach first
    summary_data = []
    if use_case and hasattr(use_case, "get_league_summary_data"):
        try:
            summary_data = use_case.get_league_summary_data()
        except Exception as e:
            logger.warning(f"Failed to get league summary from database: {e}")

    # Fallback to stats_data if database approach failed or unavailable
    if not summary_data:
        summary_data = _build_summary_from_stats_data(stats_data, query_repo)

    # Handle empty case
    if not summary_data:
        summary_data = [
            {
                "League": "No Data Available",
                "Team": "No teams found in database",
                "Games Played": 0,
                "Total Players": 0,
                "Team BA": "0.000",
                "Team OBP": "0.000",
                "Team SLG": "0.000",
                "Team OPS": "0.000",
            }
        ]

    # Create DataFrame and Excel sheet
    df = pd.DataFrame(summary_data)
    df.to_excel(writer, sheet_name="League Summary", index=False)

    # Format the sheet
    worksheet = writer.sheets["League Summary"]
    worksheet.column_dimensions["A"].width = 20  # League
    worksheet.column_dimensions["B"].width = 20  # Team
    worksheet.column_dimensions["C"].width = 15  # Games Played
    worksheet.column_dimensions["D"].width = 16  # Total Players
    worksheet.column_dimensions["E"].width = 12  # Team BA
    worksheet.column_dimensions["F"].width = 12  # Team OBP
    worksheet.column_dimensions["G"].width = 12  # Team SLG
    worksheet.column_dimensions["H"].width = 12  # Team OPS

    # Add autofilter to column headers
    worksheet.auto_filter.ref = worksheet.dimensions


def _build_summary_from_stats_data(
    stats_data: Dict[str, Any], query_repo: Optional[QueryRepository]
) -> list[Dict[str, Any]]:
    """Build league summary data from stats_data (fallback method)."""
    summary_data = []

    # Get league name from stats_data if available
    default_league = stats_data.get("league_name", "Unknown League")
    if default_league and isinstance(default_league, str):
        # Convert snake_case to Title Case (e.g., "phx_fray" -> "Phx Fray")
        default_league = " ".join(
            word.capitalize() for word in default_league.split("_")
        )

    for team_name, team_stats in stats_data.get("team_stats", {}).items():
        league_name = _get_league_for_team(team_name, query_repo, default_league)

        # Calculate team stats from player data
        players = team_stats.get("players", [])
        total_players = len(players)

        # Aggregate team totals
        total_games = team_stats.get(
            "games_played", 0
        )  # Use games_played from stats if available
        team_ba = team_stats.get("team_batting_average", 0)
        team_obp = team_stats.get("team_on_base_percentage", 0)
        team_slg = team_stats.get("team_slugging_percentage", 0)
        team_ops = team_stats.get("team_ops", 0)

        summary_data.append(
            {
                "League": league_name,
                "Team": team_name,
                "Games Played": total_games,
                "Total Players": total_players,
                "Team BA": f"{team_ba:.3f}",
                "Team OBP": f"{team_obp:.3f}",
                "Team SLG": f"{team_slg:.3f}",
                "Team OPS": f"{team_ops:.3f}",
            }
        )

    # Sort by league name, then team name
    summary_data.sort(key=lambda x: (x["League"], x["Team"]))

    return summary_data


def _create_team_sheet(
    team_name: str, team_stats: Dict[str, Any], writer: pd.ExcelWriter
) -> None:
    """Create a sheet for team statistics."""
    player_data = []

    for player_info in team_stats.get("players", []):
        player_data.append(
            {
                "Player": player_info.get(
                    "player_name", f"Player {player_info.get('player_id', 'Unknown')}"
                ),
                "PA": player_info.get("plate_appearances", 0),
                "AB": player_info.get("at_bats", 0),
                "H": player_info.get("hits", 0),
                "1B": player_info.get("singles", 0),
                "2B": player_info.get("doubles", 0),
                "3B": player_info.get("triples", 0),
                "HR": player_info.get("home_runs", 0),
                "BB": player_info.get("walks", 0),
                "SF": player_info.get("sacrifice_flies", 0),
                "RBI": player_info.get("rbis", 0),
                "R": player_info.get("runs_scored", 0),
                "BA": f"{player_info.get('batting_average', 0):.3f}",
                "OBP": f"{player_info.get('on_base_percentage', 0):.3f}",
                "SLG": f"{player_info.get('slugging_percentage', 0):.3f}",
                "OPS": f"{player_info.get('ops', 0):.3f}",
            }
        )

    if player_data:
        df = pd.DataFrame(player_data)
        # Sort by Player name alphabetically (case-insensitive)
        df = df.sort_values("Player", key=lambda x: x.str.lower(), ascending=True)

        sheet_name = f"{_abbreviate_team_name(team_name)}"  # Excel sheet names limited to 31 chars
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format the sheet
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions["A"].width = 20  # Player name
        worksheet.column_dimensions["B"].width = 8  # PA
        worksheet.column_dimensions["C"].width = 8  # AB
        worksheet.column_dimensions["D"].width = 8  # H
        worksheet.column_dimensions["E"].width = 8  # 1B
        worksheet.column_dimensions["F"].width = 8  # 2B
        worksheet.column_dimensions["G"].width = 8  # 3B
        worksheet.column_dimensions["H"].width = 8  # HR
        worksheet.column_dimensions["I"].width = 8  # BB
        worksheet.column_dimensions["J"].width = 8  # SF
        worksheet.column_dimensions["K"].width = 8  # RBI
        worksheet.column_dimensions["L"].width = 8  # R
        worksheet.column_dimensions["M"].width = 10  # BA
        worksheet.column_dimensions["N"].width = 10  # OBP
        worksheet.column_dimensions["O"].width = 10  # SLG
        worksheet.column_dimensions["P"].width = 10  # OPS

        # Add team totals row below the table with styling
        if player_data:
            totals_row = {
                "Player": "TEAM TOTALS",
                "PA": sum(player["PA"] for player in player_data),
                "AB": sum(player["AB"] for player in player_data),
                "H": sum(player["H"] for player in player_data),
                "1B": sum(player["1B"] for player in player_data),
                "2B": sum(player["2B"] for player in player_data),
                "3B": sum(player["3B"] for player in player_data),
                "HR": sum(player["HR"] for player in player_data),
                "BB": sum(player["BB"] for player in player_data),
                "SF": sum(player["SF"] for player in player_data),
                "RBI": sum(player["RBI"] for player in player_data),
                "R": sum(player["R"] for player in player_data),
                "BA": f"{team_stats.get('team_batting_average', 0):.3f}",
                "OBP": f"{team_stats.get('team_on_base_percentage', 0):.3f}",
                "SLG": f"{team_stats.get('team_slugging_percentage', 0):.3f}",
                "OPS": f"{team_stats.get('team_ops', 0):.3f}",
            }
            # Add empty separator row
            worksheet.append([])
            # Add totals row
            worksheet.append(list(totals_row.values()))
            # Apply borders to totals row to match table styling
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            totals_row_num = worksheet.max_row
            for col in range(1, 17):  # Columns A to P
                cell = worksheet.cell(row=totals_row_num, column=col)
                cell.border = thin_border

        # Add autofilter to column headers (excluding totals row)
        max_col_letter = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row - 2}"


def _create_legend_sheet(writer: pd.ExcelWriter) -> None:
    """Create a legend sheet explaining abbreviations and formulas."""
    legend_data = [
        {
            "Abbreviation": "PA",
            "Full Name": "Plate Appearances",
            "Formula": "AB + BB + SF",
        },
        {
            "Abbreviation": "AB",
            "Full Name": "At Bats",
            "Formula": "Total attempts - BB - SF",
        },
        {"Abbreviation": "H", "Full Name": "Hits", "Formula": "Total bases gained > 0"},
        {"Abbreviation": "1B", "Full Name": "Singles", "Formula": "Hits with 1 base"},
        {"Abbreviation": "2B", "Full Name": "Doubles", "Formula": "Hits with 2 bases"},
        {"Abbreviation": "3B", "Full Name": "Triples", "Formula": "Hits with 3 bases"},
        {
            "Abbreviation": "HR",
            "Full Name": "Home Runs",
            "Formula": "Hits with 4 bases",
        },
        {
            "Abbreviation": "BB",
            "Full Name": "Walks/Bases on Balls",
            "Formula": "Outcome = 'BB'",
        },
        {
            "Abbreviation": "SF",
            "Full Name": "Sacrifice Flies",
            "Formula": "Fly ball outcomes with RBI",
        },
        {
            "Abbreviation": "HRO",
            "Full Name": "Home Run Outs",
            "Formula": "Automatic outs from home run rule",
        },
        {
            "Abbreviation": "RBI",
            "Full Name": "Runs Batted In",
            "Formula": "Runs scored on play",
        },
        {
            "Abbreviation": "R",
            "Full Name": "Runs Scored",
            "Formula": "Runs scored by batter",
        },
        {"Abbreviation": "BA", "Full Name": "Batting Average", "Formula": "H / AB"},
        {
            "Abbreviation": "OBP",
            "Full Name": "On-Base Percentage",
            "Formula": "(H + BB + HBP + SF) / (AB + BB + HBP + SF)",
        },
        {
            "Abbreviation": "SLG",
            "Full Name": "Slugging Percentage",
            "Formula": "Total Bases / AB",
        },
        {
            "Abbreviation": "OPS",
            "Full Name": "On-Base Plus Slugging",
            "Formula": "OBP + SLG",
        },
    ]

    df = pd.DataFrame(legend_data)
    df.to_excel(writer, sheet_name="Legend", index=False)

    # Format the sheet
    worksheet = writer.sheets["Legend"]
    worksheet.column_dimensions["A"].width = 15  # Abbreviation
    worksheet.column_dimensions["B"].width = 25  # Full Name
    worksheet.column_dimensions["C"].width = 40  # Formula


def _create_player_summary_sheet(
    query_repo: QueryRepository, writer: pd.ExcelWriter
) -> None:
    """Create comprehensive player summary sheet with all players from all leagues/seasons, consolidated by name."""
    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    aggregated_players = {}
    player_data = []

    if isinstance(query_repo, SQLiteQueryRepository):
        with query_repo._get_connection() as conn:
            cursor = conn.cursor()
            # Get all players with their team names (but we'll aggregate by name)
            cursor.execute(
                """
                SELECT p.id, p.name, t.name as team_name
                FROM players p
                JOIN teams t ON p.team_id = t.id
                ORDER BY p.name
            """
            )

            players = cursor.fetchall()

            for player_id, player_name, team_name in players:
                stats = query_repo.get_player_stats(player_id)
                if stats:
                    if player_name not in aggregated_players:
                        aggregated_players[player_name] = {
                            "Player": player_name,
                            "PA": stats.plate_appearances,
                            "AB": stats.at_bats,
                            "H": stats.hits,
                            "1B": stats.singles,
                            "2B": stats.doubles,
                            "3B": stats.triples,
                            "HR": stats.home_runs,
                            "BB": stats.walks,
                            "SF": stats.sacrifice_flies,
                            "HRO": stats.home_run_outs,
                            "RBI": stats.rbis,
                            "R": stats.runs_scored,
                            "BA": stats.batting_average,
                            "OBP": stats.on_base_percentage,
                            "SLG": stats.slugging_percentage,
                            "OPS": stats.ops,
                        }
                    else:
                        # Sum the stats
                        aggregated_players[player_name]["PA"] += stats.plate_appearances
                        aggregated_players[player_name]["AB"] += stats.at_bats
                        aggregated_players[player_name]["H"] += stats.hits
                        aggregated_players[player_name]["1B"] += stats.singles
                        aggregated_players[player_name]["2B"] += stats.doubles
                        aggregated_players[player_name]["3B"] += stats.triples
                        aggregated_players[player_name]["HR"] += stats.home_runs
                        aggregated_players[player_name]["BB"] += stats.walks
                        aggregated_players[player_name]["SF"] += stats.sacrifice_flies
                        aggregated_players[player_name]["HRO"] += stats.home_run_outs
                        aggregated_players[player_name]["RBI"] += stats.rbis
                        aggregated_players[player_name]["R"] += stats.runs_scored

        # Recalculate derived stats after aggregation
        player_data = []
        for player in aggregated_players.values():
            ab = player["AB"]
            hits = player["H"]
            singles = player["1B"]
            doubles = player["2B"]
            triples = player["3B"]
            hr = player["HR"]

            # Recalculate averages
            player["BA"] = (
                f"{calculate_batting_average(hits, ab):.3f}" if ab > 0 else "0.000"
            )
            # Simplified OBP (BA approximation)
            player["OBP"] = player["BA"]
            player["SLG"] = (
                f"{calculate_slg(singles, doubles, triples, hr, ab):.3f}"
                if ab > 0
                else "0.000"
            )
            player[
                "OPS"
            ] = f"{calculate_ops(float(player['OBP']), float(player['SLG'])):.3f}"

            player_data.append(player)

    if player_data:
        df = pd.DataFrame(player_data)
        # Sort alphabetically by player name (case-insensitive)
        df = df.sort_values("Player", key=lambda x: x.str.lower(), ascending=True)

        df.to_excel(writer, sheet_name="Player Summary", index=False)

        # Format the sheet
        worksheet = writer.sheets["Player Summary"]
        worksheet.column_dimensions["A"].width = 20  # Player name
        worksheet.column_dimensions["B"].width = 8  # PA
        worksheet.column_dimensions["C"].width = 8  # AB
        worksheet.column_dimensions["D"].width = 8  # H
        worksheet.column_dimensions["E"].width = 8  # 1B
        worksheet.column_dimensions["F"].width = 8  # 2B
        worksheet.column_dimensions["G"].width = 8  # 3B
        worksheet.column_dimensions["H"].width = 8  # HR
        worksheet.column_dimensions["I"].width = 8  # BB
        worksheet.column_dimensions["J"].width = 8  # SF
        worksheet.column_dimensions["K"].width = 8  # HRO
        worksheet.column_dimensions["L"].width = 8  # RBI
        worksheet.column_dimensions["M"].width = 8  # R
        worksheet.column_dimensions["N"].width = 10  # BA
        worksheet.column_dimensions["O"].width = 10  # OBP
        worksheet.column_dimensions["P"].width = 10  # SLG
        worksheet.column_dimensions["Q"].width = 10  # OPS

        # Add autofilter to column headers
        worksheet.auto_filter.ref = worksheet.dimensions


def _get_seasons_for_team(
    team_name: str, query_repo: Optional[QueryRepository]
) -> list[str]:
    """Get all seasons for a team, sorted chronologically."""
    if not query_repo:
        return []
    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    if not isinstance(query_repo, SQLiteQueryRepository):
        return []
    seasons = set()
    with query_repo._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT l.season FROM leagues l
            JOIN teams t ON t.league_id = l.id
            WHERE t.name = ?
            """,
            (team_name,),
        )
        for row in cursor.fetchall():
            seasons.add(row[0])

    # Sort by year if possible
    def sort_key(s):
        parts = s.split()
        if len(parts) > 1 and parts[-1].isdigit():
            return int(parts[-1])
        return 0

    return sorted(seasons, key=sort_key)


def _get_league_for_team_season(
    team_name: str, season: str, query_repo: Optional[QueryRepository]
) -> str:
    """Get league name for a team and season."""
    if not query_repo:
        return "Unknown"
    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    if not isinstance(query_repo, SQLiteQueryRepository):
        return "Unknown"
    with query_repo._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT l.name FROM leagues l
            JOIN teams t ON t.league_id = l.id
            WHERE t.name = ? AND l.season = ?
            LIMIT 1
            """,
            (team_name, season),
        )
        row = cursor.fetchone()
        return row[0] if row else "Unknown"


def _get_team_id(
    team_name: str, season: str, query_repo: Optional[QueryRepository]
) -> Optional[int]:
    """Get team ID for team and season."""
    if not query_repo:
        return None
    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    if not isinstance(query_repo, SQLiteQueryRepository):
        return None
    with query_repo._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT t.id FROM teams t
            JOIN leagues l ON t.league_id = l.id
            WHERE t.name = ? AND l.season = ?
            LIMIT 1
            """,
            (team_name, season),
        )
        row = cursor.fetchone()
        return row[0] if row else None


def _get_all_teams(query_repo: Optional[QueryRepository]) -> list[str]:
    """Get all team names from the database."""
    if not query_repo:
        return []

    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    if not isinstance(query_repo, SQLiteQueryRepository):
        return []

    teams = []
    with query_repo._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT name FROM teams ORDER BY name")
        teams = [row[0] for row in cursor.fetchall()]

    return teams


def _get_league_for_team(
    team_name: str,
    query_repo: Optional[QueryRepository],
    default_league: str = "Unknown League",
) -> str:
    """Get league name for a team."""
    if not query_repo:
        return default_league

    from softball_statistics.repository.sqlite import SQLiteQueryRepository

    if not isinstance(query_repo, SQLiteQueryRepository):
        return default_league

    with query_repo._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT l.name FROM leagues l
            JOIN teams t ON t.league_id = l.id
            WHERE t.name = ?
            LIMIT 1
            """,
            (team_name,),
        )
        row = cursor.fetchone()
        return row[0] if row else default_league


def _create_cumulative_team_sheet(
    team_name: str, cumulative_stats: Dict[str, Any], writer: pd.ExcelWriter
) -> None:
    """Create cumulative team sheet across all seasons."""
    player_data = []

    for player_info in cumulative_stats.get("players", []):
        player_data.append(
            {
                "Player": player_info.get(
                    "player_name", f"Player {player_info.get('player_id', 'Unknown')}"
                ),
                "PA": player_info.get("plate_appearances", 0),
                "AB": player_info.get("at_bats", 0),
                "H": player_info.get("hits", 0),
                "1B": player_info.get("singles", 0),
                "2B": player_info.get("doubles", 0),
                "3B": player_info.get("triples", 0),
                "HR": player_info.get("home_runs", 0),
                "BB": player_info.get("walks", 0),
                "SF": player_info.get("sacrifice_flies", 0),
                "HRO": player_info.get("home_run_outs", 0),
                "RBI": player_info.get("rbis", 0),
                "R": player_info.get("runs_scored", 0),
                "BA": f"{player_info.get('batting_average', 0):.3f}",
                "OBP": f"{player_info.get('on_base_percentage', 0):.3f}",
                "SLG": f"{player_info.get('slugging_percentage', 0):.3f}",
                "OPS": f"{player_info.get('ops', 0):.3f}",
            }
        )

    if player_data:
        df = pd.DataFrame(player_data)
        # Sort by Player name alphabetically (case-insensitive)
        df = df.sort_values("Player", key=lambda x: x.str.lower(), ascending=True)

        sheet_name = f"{_abbreviate_team_name(team_name)} Total"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format the sheet
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = "FF0000FF"  # Blue for cumulative
        worksheet.column_dimensions["A"].width = 20  # Player name
        worksheet.column_dimensions["B"].width = 8  # PA
        worksheet.column_dimensions["C"].width = 8  # AB
        worksheet.column_dimensions["D"].width = 8  # H
        worksheet.column_dimensions["E"].width = 8  # 1B
        worksheet.column_dimensions["F"].width = 8  # 2B
        worksheet.column_dimensions["G"].width = 8  # 3B
        worksheet.column_dimensions["H"].width = 8  # HR
        worksheet.column_dimensions["I"].width = 8  # BB
        worksheet.column_dimensions["J"].width = 8  # SF
        worksheet.column_dimensions["K"].width = 8  # HRO
        worksheet.column_dimensions["L"].width = 8  # RBI
        worksheet.column_dimensions["M"].width = 8  # R
        worksheet.column_dimensions["N"].width = 10  # BA
        worksheet.column_dimensions["O"].width = 10  # OBP
        worksheet.column_dimensions["P"].width = 10  # SLG
        worksheet.column_dimensions["Q"].width = 10  # OPS

        # Add team totals row
        team_totals = cumulative_stats.get("team_totals", {})
        totals_row = {
            "Player": "TEAM TOTALS",
            "PA": sum(player["PA"] for player in player_data),
            "AB": sum(player["AB"] for player in player_data),
            "H": sum(player["H"] for player in player_data),
            "1B": sum(player["1B"] for player in player_data),
            "2B": sum(player["2B"] for player in player_data),
            "3B": sum(player["3B"] for player in player_data),
            "HR": sum(player["HR"] for player in player_data),
            "BB": sum(player["BB"] for player in player_data),
            "SF": sum(player["SF"] for player in player_data),
            "HRO": sum(player["HRO"] for player in player_data),
            "RBI": sum(player["RBI"] for player in player_data),
            "R": sum(player["R"] for player in player_data),
            "BA": f"{team_totals.get('team_batting_average', 0):.3f}",
            "OBP": f"{team_totals.get('team_on_base_percentage', 0):.3f}",
            "SLG": f"{team_totals.get('team_slugging_percentage', 0):.3f}",
            "OPS": f"{team_totals.get('team_ops', 0):.3f}",
        }
        worksheet.append([])
        worksheet.append(list(totals_row.values()))
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        totals_row_num = worksheet.max_row
        num_cols = len(totals_row)
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=totals_row_num, column=col)
            cell.border = thin_border

        # Add autofilter
        max_col_letter = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row - 2}"


def _create_season_total_sheet(
    team_name: str, season: str, season_stats: Dict[str, Any], writer: pd.ExcelWriter
) -> None:
    """Create season total sheet."""
    team_stats = season_stats.get("team_stats", {}).get(team_name, {})
    player_data = []

    for player_info in team_stats.get("players", []):
        player_data.append(
            {
                "Player": player_info.get(
                    "player_name", f"Player {player_info.get('player_id', 'Unknown')}"
                ),
                "PA": player_info.get("plate_appearances", 0),
                "AB": player_info.get("at_bats", 0),
                "H": player_info.get("hits", 0),
                "1B": player_info.get("singles", 0),
                "2B": player_info.get("doubles", 0),
                "3B": player_info.get("triples", 0),
                "HR": player_info.get("home_runs", 0),
                "BB": player_info.get("walks", 0),
                "SF": player_info.get("sacrifice_flies", 0),
                "HRO": player_info.get("home_run_outs", 0),
                "RBI": player_info.get("rbis", 0),
                "R": player_info.get("runs_scored", 0),
                "BA": f"{player_info.get('batting_average', 0):.3f}",
                "OBP": f"{player_info.get('on_base_percentage', 0):.3f}",
                "SLG": f"{player_info.get('slugging_percentage', 0):.3f}",
                "OPS": f"{player_info.get('ops', 0):.3f}",
            }
        )

    if player_data:
        df = pd.DataFrame(player_data)
        df = df.sort_values("Player", key=lambda x: x.str.lower(), ascending=True)

        sheet_name = f"{_abbreviate_team_name(team_name)} {season} Total"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = "FF00FF00"  # Green for season totals
        for col, width in [
            ("A", 20),
            ("B", 8),
            ("C", 8),
            ("D", 8),
            ("E", 8),
            ("F", 8),
            ("G", 8),
            ("H", 8),
            ("I", 8),
            ("J", 8),
            ("K", 8),
            ("L", 8),
            ("M", 8),
            ("N", 10),
            ("O", 10),
            ("P", 10),
            ("Q", 10),
        ]:
            worksheet.column_dimensions[col].width = width

        # Add totals
        team_totals = team_stats
        totals_row = {
            "Player": "TEAM TOTALS",
            "PA": sum(player["PA"] for player in player_data),
            "AB": sum(player["AB"] for player in player_data),
            "H": sum(player["H"] for player in player_data),
            "1B": sum(player["1B"] for player in player_data),
            "2B": sum(player["2B"] for player in player_data),
            "3B": sum(player["3B"] for player in player_data),
            "HR": sum(player["HR"] for player in player_data),
            "BB": sum(player["BB"] for player in player_data),
            "SF": sum(player["SF"] for player in player_data),
            "HRO": sum(player["HRO"] for player in player_data),
            "RBI": sum(player["RBI"] for player in player_data),
            "R": sum(player["R"] for player in player_data),
            "BA": f"{team_totals.get('team_batting_average', 0):.3f}",
            "OBP": f"{team_totals.get('team_on_base_percentage', 0):.3f}",
            "SLG": f"{team_totals.get('team_slugging_percentage', 0):.3f}",
            "OPS": f"{team_totals.get('team_ops', 0):.3f}",
        }
        worksheet.append([])
        worksheet.append(list(totals_row.values()))
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        totals_row_num = worksheet.max_row
        num_cols = len(totals_row)
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=totals_row_num, column=col)
            cell.border = thin_border

        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row - 2}"
        )


def _create_per_game_sheet(
    team_name: str,
    game_stat: Dict[str, Any],
    player_stats: list[Dict[str, Any]],
    writer: pd.ExcelWriter,
) -> None:
    """Create per-game sheet with player-by-player stats."""
    player_data = []

    for player_info in player_stats:
        player_data.append(
            {
                "Player": player_info.get(
                    "player_name", f"Player {player_info.get('player_id', 'Unknown')}"
                ),
                "PA": player_info.get("plate_appearances", 0),
                "AB": player_info.get("at_bats", 0),
                "H": player_info.get("hits", 0),
                "1B": player_info.get("singles", 0),
                "2B": player_info.get("doubles", 0),
                "3B": player_info.get("triples", 0),
                "HR": player_info.get("home_runs", 0),
                "BB": player_info.get("walks", 0),
                "SF": player_info.get("sacrifice_flies", 0),
                "HRO": player_info.get("home_run_outs", 0),
                "RBI": player_info.get("rbis", 0),
                "R": player_info.get("runs_scored", 0),
                "BA": f"{player_info.get('batting_average', 0):.3f}",
                "OBP": f"{player_info.get('on_base_percentage', 0):.3f}",
                "SLG": f"{player_info.get('slugging_percentage', 0):.3f}",
                "OPS": f"{player_info.get('ops', 0):.3f}",
            }
        )

    if player_data:
        df = pd.DataFrame(player_data)
        df = df.sort_values("Player", key=lambda x: x.str.lower(), ascending=True)

        sheet_name = f"{_abbreviate_team_name(team_name)} {game_stat.get('season', 'Unknown')} Game {game_stat.get('game_number', 0)}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = "FFFFFF00"  # Yellow for per-game
        for col, width in [
            ("A", 20),  # Player
            ("B", 8),  # PA
            ("C", 8),  # AB
            ("D", 8),  # H
            ("E", 8),  # 1B
            ("F", 8),  # 2B
            ("G", 8),  # 3B
            ("H", 8),  # HR
            ("I", 8),  # BB
            ("J", 8),  # SF
            ("K", 8),  # HRO
            ("L", 8),  # RBI
            ("M", 8),  # R
            ("N", 10),  # BA
            ("O", 10),  # OBP
            ("P", 10),  # SLG
            ("Q", 10),  # OPS
        ]:
            worksheet.column_dimensions[col].width = width

        # Calculate team totals from player data
        if player_data:
            totals_row = {
                "Player": "TEAM TOTALS",
                "PA": sum(player["PA"] for player in player_data),
                "AB": sum(player["AB"] for player in player_data),
                "H": sum(player["H"] for player in player_data),
                "1B": sum(player["1B"] for player in player_data),
                "2B": sum(player["2B"] for player in player_data),
                "3B": sum(player["3B"] for player in player_data),
                "HR": sum(player["HR"] for player in player_data),
                "BB": sum(player["BB"] for player in player_data),
                "SF": sum(player["SF"] for player in player_data),
                "HRO": sum(player["HRO"] for player in player_data),
                "RBI": sum(player["RBI"] for player in player_data),
                "R": sum(player["R"] for player in player_data),
                "BA": f"{calculate_batting_average(sum(player['H'] for player in player_data), sum(player['AB'] for player in player_data)):.3f}",
                "OBP": f"{calculate_batting_average(sum(player['H'] for player in player_data), sum(player['AB'] for player in player_data)):.3f}",  # Simplified
                "SLG": f"{calculate_slg(sum(player['1B'] for player in player_data), sum(player['2B'] for player in player_data), sum(player['3B'] for player in player_data), sum(player['HR'] for player in player_data), sum(player['AB'] for player in player_data)):.3f}",
                "OPS": f"{calculate_ops(calculate_batting_average(sum(player['H'] for player in player_data), sum(player['AB'] for player in player_data)), calculate_slg(sum(player['1B'] for player in player_data), sum(player['2B'] for player in player_data), sum(player['3B'] for player in player_data), sum(player['HR'] for player in player_data), sum(player['AB'] for player in player_data))):.3f}",
            }
            worksheet.append([])
            worksheet.append(list(totals_row.values()))
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            totals_row_num = worksheet.max_row
            num_cols = len(totals_row)
            for col in range(1, num_cols + 1):
                cell = worksheet.cell(row=totals_row_num, column=col)
                cell.border = thin_border

        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row - 2}"
        )


def _create_player_sheet(player_stats: PlayerStats, writer: pd.ExcelWriter) -> None:
    """Create detailed player sheet (optional)."""
    # For now, just create a simple summary
    # In the future, this could include game-by-game breakdowns
    data = [
        {
            "Statistic": "Games Played",
            "Value": "TBD",  # Would need to track this in PlayerStats
        },
        {"Statistic": "At Bats", "Value": player_stats.at_bats},
        {"Statistic": "Hits", "Value": player_stats.hits},
        {
            "Statistic": "Batting Average",
            "Value": f"{player_stats.batting_average:.3f}",
        },
        {
            "Statistic": "On Base Percentage",
            "Value": f"{player_stats.on_base_percentage:.3f}",
        },
        {
            "Statistic": "Slugging Percentage",
            "Value": f"{player_stats.slugging_percentage:.3f}",
        },
        {"Statistic": "OPS", "Value": f"{player_stats.ops:.3f}"},
    ]

    df = pd.DataFrame(data)
    sheet_name = f"Player_{player_stats.player_id}_detail"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
