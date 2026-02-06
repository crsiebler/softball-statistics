import os
import tempfile

import pytest
from openpyxl import load_workbook

from softball_statistics.exporters.excel_exporter import (
    ExcelExportError,
    _abbreviate_team_name,
    export_to_excel,
)


class TestExcelExporter:
    def test_export_basic_stats(self):
        """Test exporting basic team statistics."""
        stats_data = {
            "league_name": "Test League",
            "season": "Winter 2024",
            "team_stats": {
                "Test Team": {
                    "games_played": 5,
                    "players": [
                        {
                            "player_id": 1,
                            "player_name": "Player 1",
                            "at_bats": 20,
                            "hits": 8,
                            "singles": 5,
                            "doubles": 2,
                            "triples": 1,
                            "home_runs": 0,
                            "rbis": 6,
                            "runs_scored": 4,
                            "batting_average": 0.400,
                            "on_base_percentage": 0.450,
                            "slugging_percentage": 0.550,
                            "ops": 1.000,
                        },
                        {
                            "player_id": 2,
                            "player_name": "Player 2",
                            "at_bats": 15,
                            "hits": 6,
                            "singles": 4,
                            "doubles": 1,
                            "triples": 0,
                            "home_runs": 1,
                            "rbis": 5,
                            "runs_scored": 3,
                            "batting_average": 0.400,
                            "on_base_percentage": 0.450,
                            "slugging_percentage": 0.733,
                            "ops": 1.183,
                        },
                    ],
                    "team_batting_average": 0.400,
                    "team_on_base_percentage": 0.450,
                    "team_slugging_percentage": 0.641,
                    "team_ops": 1.091,
                }
            },
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_excel(stats_data, output_path)

            # Check that file was created
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_empty_stats(self):
        """Test exporting with no team data."""
        stats_data = {
            "league_name": "Empty League",
            "season": "Spring 2024",
            "team_stats": {},
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_excel(stats_data, output_path)

            # Check that file was created (even if empty)
            assert os.path.exists(output_path)

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_invalid_path(self):
        """Test exporting to invalid path raises error."""
        stats_data = {"team_stats": {}}

        with pytest.raises(ExcelExportError, match="Failed to export to Excel"):
            export_to_excel(stats_data, "/invalid/path/that/does/not/exist/file.xlsx")

    def test_league_column_in_summary_sheet(self):
        """Test that league name is included as first column in League Summary sheet."""
        stats_data = {
            "league_name": "phx_fray",
            "season": "Winter 2024",
            "team_stats": {
                "Test Team": {
                    "games_played": 1,
                    "players": [],
                    "team_batting_average": 0.000,
                    "team_on_base_percentage": 0.000,
                    "team_slugging_percentage": 0.000,
                    "team_ops": 0.000,
                }
            },
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_excel(stats_data, output_path)

            # Load the workbook and check the League Summary sheet
            wb = load_workbook(output_path)
            sheet = wb["League Summary"]

            # Check headers (first row)
            headers = [cell.value for cell in sheet[1]]
            assert headers[0] == "League"
            assert headers[1] == "Team"

            # Check data (second row)
            data = [cell.value for cell in sheet[2]]
            assert data[0] == "Phx Fray"  # Title case applied
            assert data[1] == "Test Team"

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_autofilter_applied_to_league_summary_sheet(self):
        """Test that autofilter is applied to League Summary sheet column headers."""
        stats_data = {
            "league_name": "test_league",
            "season": "Winter 2024",
            "team_stats": {
                "Test Team": {
                    "games_played": 1,
                    "players": [],
                    "team_batting_average": 0.000,
                    "team_on_base_percentage": 0.000,
                    "team_slugging_percentage": 0.000,
                    "team_ops": 0.000,
                }
            },
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_excel(stats_data, output_path)

            # Load the workbook and check autofilter
            wb = load_workbook(output_path)
            sheet = wb["League Summary"]

            # Check that autofilter is applied (should have a ref range)
            assert sheet.auto_filter.ref is not None
            assert sheet.auto_filter.ref != ""  # Should not be empty

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_autofilter_applied_to_team_sheet(self):
        """Test that autofilter is applied to team sheet column headers (excluding totals)."""
        stats_data = {
            "league_name": "test_league",
            "season": "Winter 2024",
            "team_stats": {
                "Cyclones": {
                    "games_played": 1,
                    "players": [
                        {
                            "player_id": 1,
                            "player_name": "Test Player",
                            "at_bats": 1,
                            "hits": 1,
                            "singles": 1,
                            "doubles": 0,
                            "triples": 0,
                            "home_runs": 0,
                            "rbis": 0,
                            "runs_scored": 0,
                            "batting_average": 1.000,
                            "on_base_percentage": 1.000,
                            "slugging_percentage": 1.000,
                            "ops": 2.000,
                        }
                    ],
                    "team_batting_average": 1.000,
                    "team_on_base_percentage": 1.000,
                    "team_slugging_percentage": 1.000,
                    "team_ops": 2.000,
                }
            },
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            export_to_excel(stats_data, output_path)

            # Load the workbook and check autofilter on abbreviated team sheet
            wb = load_workbook(output_path)
            sheet_name = _abbreviate_team_name("Cyclones")
            sheet = wb[sheet_name]

            # Check that autofilter is applied
            assert sheet.auto_filter.ref is not None
            assert sheet.auto_filter.ref != ""

            # Autofilter should cover data rows but exclude totals (last 2 rows)
            # With 1 player: rows = header(1) + player(2) + empty(3) + totals(4)
            # Filter should go to row 2 (worksheet.max_row - 2 = 4 - 2 = 2)
            expected_ref = "A1:P2"  # Assuming 16 columns (A-P)
            assert sheet.auto_filter.ref == expected_ref

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
